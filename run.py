#!/usr/bin/env python3
"""
Trend scraper CLI.

  # verify the whole pipeline offline, no network, no API keys
  python run.py --mock

  # scrape the sandbox site to check the HTML layer works
  python run.py --scrape-demo books_sandbox

  # live run
  export EBAY_CLIENT_ID=... EBAY_CLIENT_SECRET=...
  python run.py --watchlist watchlist.csv --geo IE --marketplace EBAY_GB

  # discover model names you are not tracking yet
  python run.py --discover "Coach bag" --geo IE

Output: an xlsx report plus a console table.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import scoring
from fetcher import Fetcher, RobotsDisallowed

USER_AGENT = (
    "TrendResearchBot/1.0 (+contact: you@example.com) "
    "python-requests"
)

log = logging.getLogger("trend_scraper")


WATCHLIST_FIELDS = ["search_term", "category", "brand", "model", "tier", "gender"]


def load_watchlist(path: Path) -> list[str]:
    terms: list[str] = []
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            term = (row.get("search_term") or "").strip()
            if term:
                terms.append(term)
    return terms


def append_watchlist(path: Path, rows: list[dict]) -> int:
    """Append discovered rows to the watchlist CSV, keeping its column order.
    Returns how many rows were written. Deduping is the caller's job."""
    if not rows:
        return 0
    existing = list(csv.DictReader(path.open(newline="", encoding="utf-8"))) \
        if path.exists() else []
    fieldnames = list(existing[0].keys()) if existing else WATCHLIST_FIELDS
    for field in WATCHLIST_FIELDS:              # ensure gender column exists
        if field not in fieldnames:
            fieldnames.append(field)

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in existing:
            writer.writerow(row)
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})
    return len(rows)


def load_fixtures() -> tuple[dict, dict, dict]:
    base = Path(__file__).parent / "fixtures"
    with (base / "mock_signals.json").open(encoding="utf-8") as handle:
        data = json.load(handle)
    return data["trends"], data["ebay"], data["news"]


def print_table(rows: list[scoring.Scored], limit: int) -> None:
    header = (f"{'#':>3}  {'PRODUCT':<44} {'HEAT':>5} {'BAND':>4} "
              f"{'MOM%':>7} {'LIST':>5} {'MED':>9}  VERDICT")
    print("\n" + header)
    print("-" * len(header))
    for i, row in enumerate(rows[:limit], start=1):
        print(f"{i:>3}  {row.product[:44]:<44} {row.heat_score:>5.1f} "
              f"{row.heat_band:>4} {row.momentum:>+7.1f} {row.listings:>5} "
              f"{row.median_price:>9.2f}  {row.verdict}")
    print()


def write_report(rows: list[scoring.Scored], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Signal Report"

    headers = ["Rank", "Product", "Heat Score", "Band", "Trend Momentum %",
               "Trend Level", "Active Listings", "Median Price",
               "Editorial Mentions", "Verdict"]
    ws.append(headers)
    for i, row in enumerate(rows, start=1):
        ws.append([i, row.product, row.heat_score, row.heat_band, row.momentum,
                   row.level, row.listings, row.median_price, row.mentions,
                   row.verdict])

    fill = PatternFill("solid", fgColor="1F2A44")
    for column in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=column)
        cell.fill = fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for width, column in zip([6, 46, 11, 7, 16, 11, 14, 13, 16, 44], range(1, 11)):
        ws.column_dimensions[get_column_letter(column)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    wb.save(path)
    log.info("report written: %s", path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Luxury resale trend signal collector")
    parser.add_argument("--watchlist", type=Path, default=Path("watchlist.csv"))
    parser.add_argument("--geo", default="", help='ISO code, e.g. IE, GB, US. Empty = worldwide')
    parser.add_argument("--marketplace", default="EBAY_GB")
    parser.add_argument("--out", type=Path, default=Path("signal_report.xlsx"))
    parser.add_argument("--limit", type=int, default=25, help="rows to print")
    parser.add_argument("--mock", action="store_true",
                        help="run on fixture data - no network, no keys")
    parser.add_argument("--scrape-demo", metavar="PROFILE",
                        help="scrape a sandbox site profile to test the HTML layer")
    parser.add_argument("--discover", metavar="SEED",
                        help="list rising related queries for a seed term")
    parser.add_argument("--auto-discover", action="store_true",
                        help="find trending products (women's + men's) and add "
                             "the new ones to the watchlist before scoring")
    parser.add_argument("--no-ebay", action="store_true")
    parser.add_argument("--no-news", action="store_true")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)-7s %(message)s",
    )

    # ---- HTML layer smoke test -------------------------------------
    if args.scrape_demo:
        from sources import generic_html
        fetcher = Fetcher(user_agent=USER_AGENT)
        try:
            records = generic_html.scrape(args.scrape_demo, fetcher, max_pages=2)
        except RobotsDisallowed as exc:
            log.error("%s", exc)
            return 2
        for record in records[:10]:
            print(record)
        print(f"\n{len(records)} records scraped.")
        return 0

    # ---- discovery mode --------------------------------------------
    if args.discover:
        from sources import google_trends
        for query in google_trends.rising_queries(args.discover, geo=args.geo):
            print(query)
        return 0

    # ---- signal collection -----------------------------------------
    if args.mock:
        log.info("MOCK MODE - fixture data, no network calls")
        trends, ebay, news = load_fixtures()
    else:
        if not args.watchlist.exists():
            log.error("watchlist not found: %s", args.watchlist)
            return 1
        terms = load_watchlist(args.watchlist)
        log.info("loaded %d watchlist terms", len(terms))

        if args.auto_discover:
            from sources import autodiscover
            new_rows = autodiscover.expand(terms, geo=args.geo)
            written = append_watchlist(args.watchlist, new_rows)
            if written:
                terms = load_watchlist(args.watchlist)
                log.info("watchlist now %d terms after discovery", len(terms))

        from sources import google_trends
        trends = google_trends.collect(terms, geo=args.geo)

        ebay = {}
        if not args.no_ebay:
            from sources import ebay_api
            ebay = ebay_api.collect(terms, marketplace=args.marketplace)

        news = {}
        if not args.no_news:
            from sources import rss_news
            news = rss_news.collect(terms)

    rows = scoring.blend(trends, ebay, news)
    print_table(rows, args.limit)
    write_report(rows, args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
